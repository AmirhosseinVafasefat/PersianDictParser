from openpyxl import load_workbook
import re
from docx import Document
import argparse


def main():
    docx_doc, xlsx_doc, sheet_name, final_xlsx_doc, word_column, meaning_column, synonym_column, examples_column = arg_parser()

    word_doc = Document(docx_doc)
    excel_book = load_workbook(xlsx_doc)
    excel_sheet = excel_book[sheet_name]
    row = find_empty_row(excel_sheet)

    for paragraph in word_doc.paragraphs:
        meanings = []
        parse_meanings(paragraph.text, meanings)
        
        for meaning_counter, text in enumerate(meanings):
            if meaning_counter == 0:
                word, meaning, synonym, examples = parser_first_part(text)
                if len(meanings) > 1:
                    meaning = '1.' + meaning
                    if synonym:
                        synonym = '1.' + synonym
                    if examples:
                        examples = '1.' + examples
            else:
                returned_meaning, returned_synonym, returned_examples = parse_remainder(text)
                meaning = meaning + f'\n{meaning_counter + 1}.' + returned_meaning
                if returned_synonym:
                    if synonym:
                        synonym = synonym + f'\n{meaning_counter + 1}.' + returned_synonym
                    else:
                        synonym = f'\n{meaning_counter + 1}.' + returned_synonym
                if returned_examples:
                    if examples:
                        examples = examples + f'\n{meaning_counter + 1}.' + returned_examples
                    else:
                        examples = f'\n{meaning_counter + 1}.' + returned_examples

        fill_row(excel_sheet, row, word_column, meaning_column, synonym_column, examples_column, word, meaning, synonym, examples)
        row += 1

    excel_book.save(final_xlsx_doc)


def find_empty_row(sheet):
    '''finds the next empty row.'''
    for row in range(1, sheet.max_row + 1):
        if all([cell.value is None for cell in sheet[row]]):
            return row
    return sheet.max_row + 1

def parse_meanings(text, meanings):
    '''recursivly parses the text to get different meanings (divided by "||") and adds them to the list named meanings.'''

    matches = re.search(r"^(.+?) *\|\| *(.+)$", text.strip())
    if not matches:
        meanings.append(text)
    if matches:
        text_, remainder = matches.groups()
        meanings.append(text_)
        parse_meanings(remainder, meanings)

def parser_first_part(text):
    '''parses text and enters them into the xlsx file'''

    matches = re.search(r"^(.+?)\. *(.+?) *(?:\(مترادف *: *(.+)\))? *(?::) *(«.+)$", text.strip())
    if matches:
        word, meaning, synonym, examples = matches.groups()
        return word, meaning, synonym, examples
        
    elif not matches:
        matches = re.search(r"^(.+?)\. *(.+)$", text.strip())
        if matches:
            word, meaning = matches.groups()
            return word, meaning, None, None


def parse_remainder (text):
    '''parses text and enters them into the xlsx file'''

    matches = re.search(r"^(.+?) *(?:\(مترادف *: *(.+)\))? *(?::) *(«.+)", text.strip())
    if matches:
        meaning, synonym, examples = matches.groups()
        return meaning, synonym, examples
    else:
        return text, None, None
    
def fill_row(excel_sheet, row, word_column, meaning_column, synonym_column, examples_column, word, meaning, synonym, examples):
    excel_sheet[f'{word_column}{row}'].value = word
    excel_sheet[f'{meaning_column}{row}'].value = meaning
    excel_sheet[f'{synonym_column}{row}'].value = synonym
    excel_sheet[f'{examples_column}{row}'].value = examples

def arg_parser():
    '''parses the command line arguments.'''

    commandline_parser = argparse.ArgumentParser(description='This is a program for parsing words, meanings, synonyms, and examples from a Persian Dictionary and puting it in an excel sheet.')
    commandline_parser.add_argument('docx_document', metavar='docx_document', type=str, help='enter the name for the docx document.')
    commandline_parser.add_argument('xlsx_document', metavar='xlsx_document', type=str, help='enter the name for the xlsx document.')
    commandline_parser.add_argument('sheet_name', metavar='sheet_name', type=str, help='enter the sheet name in the xlsx document.')
    commandline_parser.add_argument('final_xlsx_document', metavar='final_xlsx_document', type=str, help='enter the name for the xlsx document you want to save into.')
    commandline_parser.add_argument('column_for_word', metavar='column_for_word', type=str, help='enter the column for the words.')
    commandline_parser.add_argument('column_for_meaning', metavar='column_for_meaning', type=str, help='enter the column for the meanings.')
    commandline_parser.add_argument('column_for_synonym', metavar='column_for_synonym', type=str, help='enter the column for the synonyms.')
    commandline_parser.add_argument('column_for_examples', metavar='column_for_examples', type=str, help='enter the column for the examples.')

    args = commandline_parser.parse_args()
    
    return args.docx_document, args.xlsx_document, args.sheet_name, args.final_xlsx_document, args.column_for_word, args.column_for_meaning, args.column_for_synonym, args.column_for_examples


if __name__ == "__main__":
    main()




